from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import logging
from logging.handlers import RotatingFileHandler

app = Flask(__name__)

# Configure logging
if not os.path.exists('logs'):
    os.mkdir('logs')
file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240, backupCount=10)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
))
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)
app.logger.info('Application startup')

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///site.db')
if app.config['SQLALCHEMY_DATABASE_URI'].startswith("postgres://"):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)

class Block(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    floors = db.Column(db.String(200), nullable=False)  # Store as JSON string
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def init_db():
    try:
        with app.app_context():
            db.create_all()
            # Create admin user if it doesn't exist
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                admin = User(
                    username='admin',
                    password=bcrypt.generate_password_hash('admin123').decode('utf-8'),
                    is_admin=True
                )
                db.session.add(admin)
                db.session.commit()
                app.logger.info('Admin user created successfully')
    except Exception as e:
        app.logger.error(f'Error initializing database: {str(e)}')
        raise

# Initialize database
init_db()

# Error handlers
@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    app.logger.error(f'Server Error: {error}')
    return render_template('error.html', error="An internal error occurred. Please try again later."), 500

@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', error="Page not found."), 404

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            user = User.query.filter_by(username=username).first()
            
            if user and bcrypt.check_password_hash(user.password, password):
                login_user(user)
                app.logger.info(f'User {username} logged in successfully')
                return redirect(url_for('dashboard'))
            flash('Invalid username or password')
            app.logger.warning(f'Failed login attempt for user {username}')
        except Exception as e:
            app.logger.error(f'Error during login: {str(e)}')
            flash('An error occurred during login. Please try again.')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            
            if not username or not password:
                flash('Username and password are required')
                return redirect(url_for('register'))
            
            if User.query.filter_by(username=username).first():
                flash('Username already exists')
                return redirect(url_for('register'))
            
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            new_user = User(username=username, password=hashed_password)
            db.session.add(new_user)
            db.session.commit()
            
            app.logger.info(f'New user registered: {username}')
            flash('Registration successful! Please login.')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error during registration: {str(e)}')
            flash('An error occurred during registration. Please try again.')
    return render_template('register.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated and current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username, is_admin=True).first()
        
        if user and bcrypt.check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Invalid admin credentials')
    return render_template('admin_login.html')

@app.route('/dashboard')
@login_required
def dashboard():
    blocks = Block.query.all()
    block_data = {}
    for block in blocks:
        block_data[block.name] = json.loads(block.floors)
    return render_template('dashboard.html', block_data=block_data)

@app.route('/update-block', methods=['POST'])
@login_required
def update_block():
    data = request.get_json()
    block_name = data.get('block')
    floors = data.get('floors')
    
    block = Block.query.filter_by(name=block_name).first()
    if block:
        block.floors = json.dumps(floors)
    else:
        block = Block(name=block_name, floors=json.dumps(floors))
        db.session.add(block)
    
    db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/delete-block', methods=['POST'])
@login_required
def delete_block():
    data = request.get_json()
    block_name = data.get('block')
    
    block = Block.query.filter_by(name=block_name).first()
    if block:
        db.session.delete(block)
        db.session.commit()
    
    return jsonify({'status': 'success'})

@app.route('/report')
@login_required
def report():
    blocks = Block.query.all()
    block_summary = []
    
    for block in blocks:
        floors = json.loads(block.floors)
        completed_floors = sum(1 for f in floors if f)
        percent = (completed_floors / len(floors)) * 100 if floors else 0
        
        block_summary.append({
            'block': block.name,
            'completed_floors': completed_floors,
            'percent': percent
        })
    
    return render_template('report.html', block_summary=block_summary)

@app.route('/download')
@login_required
def download():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Field Service Progress"
    
    # Headers
    headers = ['Block', 'Floor 1', 'Floor 2', 'Floor 3', 'Floor 4', 'Floor 5', 'Floor 6', 'Floor 7', 'Completion %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    blocks = Block.query.all()
    for row, block in enumerate(blocks, 2):
        floors = json.loads(block.floors)
        completed = sum(1 for f in floors if f)
        percent = (completed / len(floors)) * 100 if floors else 0
        
        ws.cell(row=row, column=1, value=block.name)
        for col, status in enumerate(floors, 2):
            cell = ws.cell(row=row, column=col)
            cell.value = '✓' if status else '✗'
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=9, value=f"{percent:.1f}%")
    
    # Save file
    filename = f"field_service_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    return send_file(filename, as_attachment=True)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
